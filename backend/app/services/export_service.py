import os
import logging
from pathlib import Path
from app.models.models import Analysis, Company, Financial

logger = logging.getLogger("investorgpt.export_service")

class ExportService:
    """Generates PDF and Excel reports for a completed analysis."""

    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_pdf_report(self, analysis: Analysis, company: Company, financials: list[Financial]) -> str:
        """Generate PDF file using ReportLab and return path."""
        pdf_path = self.output_dir / f"report_{analysis.id}.pdf"
        logger.info(f"Generating PDF report at {pdf_path}")

        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            # Custom styles
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=24,
                leading=28,
                textColor=colors.HexColor('#8b5cf6') # Purple accent
            )

            # Header
            story.append(Paragraph(f"InvestorGPT Research Report", title_style))
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"Company: <b>{company.name} ({company.ticker})</b>", styles['Normal']))
            story.append(Paragraph(f"Exchange: {company.exchange} | Country: {company.country}", styles['Normal']))
            story.append(Paragraph(f"Sector: {company.sector} | Industry: {company.industry}", styles['Normal']))
            story.append(Spacer(1, 15))

            # Recommendation
            story.append(Paragraph("<b>Committee Recommendation</b>", styles['Heading2']))
            story.append(Paragraph(f"Consensus: <b>{analysis.recommendation}</b>", styles['Normal']))
            story.append(Paragraph(f"Confidence Score: {round((analysis.confidence or 0)*100)}%" if hasattr(analysis, "confidence") and analysis.confidence else "Confidence: N/A", styles['Normal']))
            story.append(Spacer(1, 15))

            # Resolve currency symbol for PDF
            currency_code = (company.currency or "USD").upper().strip()
            pdf_currency = "Rs. " if currency_code == "INR" else "€" if currency_code == "EUR" else "£" if currency_code == "GBP" else "¥" if currency_code == "JPY" else "$"

            # Financial Table
            story.append(Paragraph("<b>Normalized Financials</b>", styles['Heading2']))
            
            data = [["Metric", "Fiscal Period", "Value", "Source"]]
            for f in financials[:15]:  # Limit to first 15 for spacing
                val_formatted = str(f.value)
                if f.value is not None:
                    if f.metric_name == "f_score":
                        val_formatted = f"{int(f.value)}/9"
                    elif f.metric_name == "z_score" or f.metric_name == "sentiment_score":
                        val_formatted = f"{float(f.value):.2f}"
                    elif abs(float(f.value)) > 1e6:
                        val_formatted = f"{pdf_currency}{float(f.value) / 1e9:.2f}B"
                    else:
                        val_formatted = f"{pdf_currency}{float(f.value):.2f}"
                data.append([
                    f.metric_name.replace("_", " ").title(),
                    f.fiscal_period or "N/A",
                    val_formatted,
                    f.source
                ])

            t = Table(data, hAlign='LEFT')
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f2937')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#374151')),
                ('FONTSIZE', (0,0), (-1,-1), 9),
            ]))
            story.append(t)

            doc.build(story)
            return str(pdf_path)

        except Exception as e:
            logger.error(f"Failed to generate ReportLab PDF: {e}. Creating fallback txt file.")
            # Fallback to plain text masquerading as PDF or just standard txt output
            fallback_path = self.output_dir / f"report_{analysis.id}.txt"
            with open(fallback_path, "w", encoding="utf-8") as f:
                f.write(f"InvestorGPT Research Report\n")
                f.write(f"===========================\n")
                f.write(f"Company: {company.name} ({company.ticker})\n")
                f.write(f"Recommendation: {analysis.recommendation}\n")
                f.write(f"Confidence: {analysis.confidence}\n\n")
                for fin in financials:
                    f.write(f"{fin.metric_name}: {fin.value} ({fin.fiscal_period})\n")
            return str(fallback_path)

    def generate_excel_report(self, analysis: Analysis, company: Company, financials: list[Financial]) -> str:
        """Generate Excel file using openpyxl and return path."""
        xlsx_path = self.output_dir / f"report_{analysis.id}.xlsx"
        logger.info(f"Generating Excel report at {xlsx_path}")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Analysis"

            # Header Styling
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            # Title
            ws["A1"] = f"{company.name} ({company.ticker}) Financial Analysis"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"] = f"Recommendation: {analysis.recommendation} | Confidence: {analysis.confidence}"

            headers = ["Metric Name", "Fiscal Period", "Value", "Source", "Confidence"]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")

            row_idx = 5
            for f in financials:
                ws.cell(row=row_idx, column=1, value=f.metric_name)
                ws.cell(row=row_idx, column=2, value=f.fiscal_period)
                ws.cell(row=row_idx, column=3, value=float(f.value) if f.value is not None else None)
                ws.cell(row=row_idx, column=4, value=f.source)
                ws.cell(row=row_idx, column=5, value=float(f.confidence))
                row_idx += 1

            wb.save(str(xlsx_path))
            return str(xlsx_path)

        except Exception as e:
            logger.error(f"Failed to generate Excel sheet: {e}")
            # Fallback mock file
            with open(xlsx_path, "w", encoding="utf-8") as f:
                f.write("CSV representation\n")
                for fin in financials:
                    f.write(f"{fin.metric_name},{fin.fiscal_period},{fin.value}\n")
            return str(xlsx_path)

    def generate_comparison_report_pdf(self, comparison_data: list[dict]) -> str:
        """Generate a beautiful comparison landscape PDF report."""
        pdf_path = self.output_dir / "comparison_report.pdf"
        logger.info(f"Generating PDF comparison report at {pdf_path}")
        
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            doc = SimpleDocTemplate(str(pdf_path), pagesize=landscape(letter))
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CompTitle',
                parent=styles['Heading1'],
                fontSize=20,
                leading=24,
                textColor=colors.HexColor('#8b5cf6')
            )
            
            story.append(Paragraph("InvestorGPT - Side-by-Side Stock Comparison", title_style))
            story.append(Spacer(1, 12))
            
            headers = ["Metric"] + [comp["ticker"] for comp in comparison_data]
            rows_def = [
                ("Company Name", "name"),
                ("Stock Price", "price"),
                ("DCF Fair Value", "fair_value"),
                ("Piotroski F-Score", "f_score"),
                ("Altman Z-Score", "z_score"),
                ("P/E Ratio", "pe"),
                ("RSI (14)", "rsi"),
                ("News Sentiment", "sentiment"),
                ("Risk Level", "risk_level"),
                ("Annual Revenue", "revenue"),
                ("Gross Margin", "gross_margin"),
                ("Net Margin", "net_margin"),
                ("Operating Cash Flow", "operating_cash_flow"),
                ("Capital Expenditures", "capital_expenditures")
            ]
            
            table_data = [headers]
            for label, field in rows_def:
                row = [label]
                for comp in comparison_data:
                    val = comp.get(field)
                    currency = comp.get("currency", "USD")
                    symbol = "Rs. " if currency == "INR" else "€" if currency == "EUR" else "£" if currency == "GBP" else "¥" if currency == "JPY" else "$"
                    
                    if val is None:
                        row.append("N/A")
                    elif field == "name":
                        row.append(str(val))
                    elif field == "price" or field == "fair_value":
                        row.append(f"{symbol}{float(val):.2f}")
                    elif field == "f_score":
                        row.append(f"{int(val)}/9")
                    elif field == "z_score":
                        row.append(f"{float(val):.2f}")
                    elif field == "pe":
                        row.append(f"{float(val):.1f}")
                    elif field == "rsi":
                        row.append(f"{float(val):.1f}")
                    elif field in ["sentiment", "risk_level"]:
                        row.append(str(val))
                    elif field in ["revenue", "operating_cash_flow"]:
                        row.append(f"{symbol}{float(val) / 1e9:.2f}B")
                    elif field == "capital_expenditures":
                        row.append(f"-{symbol}{abs(float(val)) / 1e9:.2f}B")
                    elif field in ["gross_margin", "net_margin"]:
                        row.append(f"{float(val) * 100:.1f}%")
                    else:
                        row.append(str(val))
                table_data.append(row)
                
            col_width = 550 / max(1, len(headers) - 1)
            t = Table(table_data, colWidths=[150] + [col_width] * (len(headers) - 1))
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f2937')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#374151')),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            story.append(t)
            doc.build(story)
            return str(pdf_path)
        except Exception as e:
            logger.error(f"Failed to generate comparison PDF: {e}")
            raise e

    def generate_comparison_report_excel(self, comparison_data: list[dict]) -> str:
        """Generate a clean side-by-side comparison Excel sheet."""
        xlsx_path = self.output_dir / "comparison_report.xlsx"
        logger.info(f"Generating Excel comparison report at {xlsx_path}")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparison"
            
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            
            ws["A1"] = "InvestorGPT - Side-by-Side Stock Comparison"
            ws["A1"].font = Font(size=14, bold=True)
            
            headers = ["Metric"] + [comp["ticker"] for comp in comparison_data]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")
                
            rows_def = [
                ("Company Name", "name"),
                ("Stock Price", "price"),
                ("DCF Fair Value", "fair_value"),
                ("Piotroski F-Score", "f_score"),
                ("Altman Z-Score", "z_score"),
                ("P/E Ratio", "pe"),
                ("RSI (14)", "rsi"),
                ("News Sentiment", "sentiment"),
                ("Risk Level", "risk_level"),
                ("Annual Revenue", "revenue"),
                ("Gross Margin", "gross_margin"),
                ("Net Margin", "net_margin"),
                ("Operating Cash Flow", "operating_cash_flow"),
                ("Capital Expenditures", "capital_expenditures")
            ]
            
            for row_idx, (label, field) in enumerate(rows_def, 4):
                ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                for col_idx, comp in enumerate(comparison_data, 2):
                    val = comp.get(field)
                    ws.cell(row=row_idx, column=col_idx, value=val)
                    
            wb.save(str(xlsx_path))
            return str(xlsx_path)
        except Exception as e:
            logger.error(f"Failed to generate comparison Excel: {e}")
            raise e

