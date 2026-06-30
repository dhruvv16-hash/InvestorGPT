import json
import logging
import io
import httpx
from datetime import datetime, timezone
from typing import Any, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database.db import get_db
from app.models.models import FinancialModel, Company, User
from app.dependencies import get_current_user
from app.providers.market.yahoo_provider import YahooProvider
from app.engines.macro_engine import MacroEngine
from app.engines.valuation.reviewer_engine import ReviewerEngine
from app.engines.valuation.modeling_engine import (
    calculate_wacc,
    run_three_statement_model,
    run_reverse_dcf,
    generate_monte_carlo,
    run_consensus_intrinsic_value,
    generate_automatic_assumptions,
    calculate_historical_valuation
)
from app.config import settings

logger = logging.getLogger("investorgpt.routes_modeling")
router = APIRouter(prefix="/modeling", tags=["Financial Modeling"])

# Pydantic Schemas
class ModelSaveRequest(BaseModel):
    ticker: str
    user_id: str
    name: str
    assumptions: dict[str, Any]

class AIChatRequest(BaseModel):
    ticker: str
    user_id: str
    prompt: str
    current_assumptions: dict[str, Any]

# Helper to load historical data
async def fetch_historical_financials(ticker: str) -> dict[str, Any]:
    provider = YahooProvider()
    return await provider.get_financial_statements(ticker)

# 1. GET Workspace list
@router.get("/workspace")
def get_workspace(
    ticker: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    models = db.query(FinancialModel).filter(
        FinancialModel.ticker == ticker.upper(),
        FinancialModel.user_id == current_user.id
    ).order_by(FinancialModel.timestamp.desc()).all()
    
    results = []
    for m in models:
        results.append({
            "id": m.id,
            "name": m.name,
            "timestamp": m.timestamp.isoformat(),
            "assumptions": m.assumptions
        })
    return {"workspace": results}

# 2. GET Model Projections
@router.get("/model/{model_id}")
async def get_model_projections(
    model_id: str,
    ticker: str = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Retrieve historical data
    try:
        hist_financials = await fetch_historical_financials(ticker)
    except Exception as e:
        logger.error(f"Failed to fetch historical statements for {ticker}: {e}")
        raise HTTPException(status_code=400, detail=f"Could not load financial data for {ticker}")

    # Load defaults
    provider = YahooProvider()
    currency = "USD"
    try:
        price_profile = await provider.get_price(ticker)
        current_price = price_profile.get("price", 100.0)
        market_cap = price_profile.get("market_cap", 1e10)
        beta = price_profile.get("beta", 1.0)
        shares = price_profile.get("shares_outstanding", 1e8)
        currency = price_profile.get("currency", "USD")
    except Exception:
        current_price = 100.0
        market_cap = 1e10
        beta = 1.0
        shares = 1e8
        currency = "USD"

    # Extract historical debt & cash
    hist_years = sorted([y for y in hist_financials.get("revenue", {}).keys() if y.isdigit()])
    base_year = hist_years[-1] if hist_years else "2025"
    
    total_debt = hist_financials.get("long_term_debt", {}).get(base_year, 0.0)
    cash = hist_financials.get("cash", {}).get(base_year, 0.0)
    net_debt = total_debt - cash

    # Compute default WACC using CAPM
    default_wacc_val = calculate_wacc(market_cap=market_cap, total_debt=total_debt, beta=beta)

    # Default overrides/assumptions
    assumptions = {
        "wacc": default_wacc_val,
        "shares_outstanding": shares,
        "terminal_growth": 0.025
    }

    # If it is a specific model ID, query database overrides
    model_name = "Default Base"
    if model_id != "default":
        db_model = db.query(FinancialModel).filter(FinancialModel.id == model_id).first()
        if db_model:
            assumptions.update(db_model.assumptions)
            model_name = db_model.name
    
    # Fetch macroeconomic indicators (AI Economist)
    macro_engine = MacroEngine()
    macro_data = await macro_engine.get_macro_indicators("USA")

    # Apply calculations
    model_results = run_three_statement_model(hist_financials, assumptions, macro_data=macro_data)
    
    # Calculate historical valuation metrics
    hist_val_data = calculate_historical_valuation(hist_financials, current_price, shares)
    
    # Calculate reverse DCF growth rate
    fcf_base = model_results["projected_fcfs"][0] if model_results["projected_fcfs"] else 1e7
    market_growth = run_reverse_dcf(
        current_price=current_price,
        fcf_base=fcf_base,
        wacc=model_results["assumptions"]["wacc"],
        terminal_growth=model_results["assumptions"]["terminal_growth"],
        years=10,
        net_debt=net_debt,
        shares_outstanding=shares
    )

    # Run Monte Carlo (10k simulations)
    monte_carlo_res = generate_monte_carlo(
        fcf_base=fcf_base,
        base_growth=model_results["assumptions"]["revenue_growth"],
        wacc=model_results["assumptions"]["wacc"],
        terminal_growth=model_results["assumptions"]["terminal_growth"],
        years=10,
        net_debt=net_debt,
        shares_outstanding=shares,
        simulations=10000
    )

    # WACC vs Growth Sensitivity Matrix (2D)
    wacc_base = model_results["assumptions"]["wacc"]
    growth_base = model_results["assumptions"]["revenue_growth"]
    
    wacc_steps = [wacc_base + adj for adj in [-0.02, -0.01, 0.0, 0.01, 0.02]]
    growth_steps = [growth_base + adj for adj in [-0.02, -0.01, 0.0, 0.01, 0.02]]
    
    matrix = []
    for w in wacc_steps:
        row = []
        for g in growth_steps:
            try:
                # Run simplified quick dcf valuation for cell
                fcfs = []
                f = fcf_base
                for _ in range(10):
                    f = f * (1 + g)
                    fcfs.append(f)
                disc_fcfs = [fcf / ((1 + w) ** (t + 1)) for t, fcf in enumerate(fcfs)]
                tv = (fcfs[-1] * (1 + model_results["assumptions"]["terminal_growth"])) / (w - model_results["assumptions"]["terminal_growth"]) if w > model_results["assumptions"]["terminal_growth"] else fcfs[-1] * 20.0
                disc_tv = tv / ((1 + w) ** 10)
                cell_val = ((sum(disc_fcfs) + disc_tv) - net_debt) / shares
                row.append(max(0.0, cell_val))
            except Exception:
                row.append(0.0)
        matrix.append(row)

    # Relative Multiples Comparisons
    relative_peers = [
        {"metric": "P/E Ratio", "company": current_price / (model_results["model"][base_year]["eps"] or 1.0), "AMD": 45.2, "INTC": 18.5, "Industry": 28.6},
        {"metric": "EV/EBITDA", "company": model_results["enterprise_value"] / (model_results["model"][base_year]["ebitda"] or 1.0), "AMD": 32.1, "INTC": 12.4, "Industry": 22.1},
        {"metric": "EV/Sales", "company": model_results["enterprise_value"] / (model_results["model"][base_year]["revenue"] or 1.0), "AMD": 8.5, "INTC": 3.2, "Industry": 6.1}
    ]

    # Weighted Consensus Intrinsic Value
    peg_val = current_price * 0.95
    hist_val = hist_val_data["current"]["pe"] * current_price / (hist_val_data["averages"]["pe"] or 1.0) if hist_val_data["averages"]["pe"] > 0 else current_price
    
    consensus_res = run_consensus_intrinsic_value(
        dcf_val=model_results["intrinsic_value"],
        comparable_val=current_price * 1.02,  # comps mock
        reverse_dcf_val=current_price * 0.98,
        peg_val=peg_val,
        historical_val=hist_val,
        residual_income_val=model_results["intrinsic_value"] * 0.96,
        ev_ebitda_val=current_price * 1.01,
        industry_multiple_val=current_price * 1.03
    )

    margin_of_safety = 0.0
    if consensus_res["intrinsic_value"] > 0:
        margin_of_safety = (consensus_res["intrinsic_value"] - current_price) / consensus_res["intrinsic_value"]

    # Generate Fair Value Tracker from saved models history
    tracker_models = db.query(FinancialModel).filter(
        FinancialModel.ticker == ticker.upper(),
        FinancialModel.user_id == current_user.id
    ).order_by(FinancialModel.timestamp.asc()).all()
    
    tracker_timeline = []
    
    # Baseline node
    default_results = run_three_statement_model(hist_financials, {}, macro_data=macro_data)
    default_dcf = default_results["intrinsic_value"]
    default_consensus = run_consensus_intrinsic_value(
        dcf_val=default_dcf,
        comparable_val=current_price * 1.02,
        reverse_dcf_val=current_price * 0.98,
        peg_val=peg_val,
        historical_val=hist_val,
        residual_income_val=default_dcf * 0.96,
        ev_ebitda_val=current_price * 1.01,
        industry_multiple_val=current_price * 1.03
    )["intrinsic_value"]
    
    tracker_timeline.append({
        "name": "Default Base",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "intrinsic_value": default_consensus,
        "margin_of_safety": (default_consensus - current_price) / default_consensus if default_consensus > 0 else 0.0
    })
    
    for m in tracker_models:
        m_results = run_three_statement_model(hist_financials, m.assumptions, macro_data=macro_data)
        m_dcf = m_results["intrinsic_value"]
        m_consensus = run_consensus_intrinsic_value(
            dcf_val=m_dcf,
            comparable_val=current_price * 1.02,
            reverse_dcf_val=current_price * 0.98,
            peg_val=peg_val,
            historical_val=hist_val,
            residual_income_val=m_dcf * 0.96,
            ev_ebitda_val=current_price * 1.01,
            industry_multiple_val=current_price * 1.03
        )["intrinsic_value"]
        
        tracker_timeline.append({
            "name": m.name,
            "timestamp": m.timestamp.isoformat(),
            "intrinsic_value": m_consensus,
            "margin_of_safety": (m_consensus - current_price) / m_consensus if m_consensus > 0 else 0.0
        })

    # Run ReviewerEngine audits (Reality Checker, Peer & Historical Validators)
    reviewer = ReviewerEngine()
    reality_warnings = reviewer.run_reality_checker(model_results)
    
    # Calculate historical revenue CAGR
    rev_hist = hist_financials.get("revenue", {})
    hist_years_sorted = sorted([y for y in rev_hist.keys() if y.isdigit()])
    historical_cagr = 0.0
    if len(hist_years_sorted) > 1:
        rev_vals = [rev_hist[y] for y in hist_years_sorted if rev_hist[y] > 0]
        if len(rev_vals) > 1:
            historical_cagr = (rev_vals[-1] / rev_vals[0]) ** (1 / (len(rev_vals) - 1)) - 1
            
    peer_warnings = await reviewer.run_peer_historical_validator(
        ticker=ticker,
        industry=price_profile.get("industry", "technology") if "price_profile" in locals() else "technology",
        model_data=model_results,
        historical_cagr=historical_cagr
    )
    audit = reviewer.audit_model(model_results, reality_warnings, peer_warnings)

    return {
        "model_name": model_name,
        "ticker": ticker.upper(),
        "currency": currency,
        "current_price": current_price,
        "intrinsic_value": consensus_res["intrinsic_value"],
        "margin_of_safety": margin_of_safety,
        "dcf_value": model_results["intrinsic_value"],
        "three_statement": model_results["model"],
        "hist_years": model_results["hist_years"],
        "proj_years": model_results["proj_years"],
        "projected_fcfs": model_results["projected_fcfs"],
        "discounted_fcfs": model_results["discounted_fcfs"],
        "discounted_tv": model_results["discounted_tv"],
        "enterprise_value": model_results["enterprise_value"],
        "net_debt": model_results["net_debt"],
        "assumptions": model_results["assumptions"],
        "auto_data": model_results["auto_data"],
        "health_score": model_results.get("health_score"),
        "audit": audit,
        "historical_valuation": hist_val_data,
        "reverse_dcf_growth": market_growth,
        "monte_carlo": monte_carlo_res,
        "sensitivity": {
            "wacc_labels": [f"{w*100:.1f}%" for w in wacc_steps],
            "growth_labels": [f"{g*100:.1f}%" for g in growth_steps],
            "matrix": matrix
        },
        "relative_valuation": relative_peers,
        "tracker_timeline": tracker_timeline,
        "consensus_details": consensus_res
    }

# 3. POST Save Model
@router.post("/save")
def save_model(
    req: ModelSaveRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    existing = db.query(FinancialModel).filter(
        FinancialModel.ticker == req.ticker.upper(),
        FinancialModel.user_id == current_user.id,
        FinancialModel.name == req.name
    ).first()
    
    if existing:
        existing.assumptions = req.assumptions
        existing.timestamp = datetime.now(timezone.utc)
        db.commit()
        return {"status": "success", "id": existing.id, "message": "Model updated successfully"}
    
    new_m = FinancialModel(
        ticker=req.ticker.upper(),
        user_id=current_user.id,
        name=req.name,
        assumptions=req.assumptions
    )
    db.add(new_m)
    db.commit()
    db.refresh(new_m)
    return {"status": "success", "id": new_m.id, "message": "Model saved successfully"}

# 4. POST Conversational Chat Editor
@router.post("/chat")
async def chat_modeling_assistant(req: AIChatRequest):
    system_prompt = (
        "You are an expert investment banking analyst. Your job is to translate conversational requests "
        "into quantitative financial modeling assumptions. Match the prompt to one or more of these keys:\n"
        "- revenue_growth (float, e.g. 0.08 for 8%)\n"
        "- gross_margin (float, e.g. 0.35 for 35%)\n"
        "- ebit_margin (float, e.g. 0.18 for 18%)\n"
        "- tax_rate (float)\n"
        "- capex_pct (float, CapEx as % of Revenue)\n"
        "- wacc (float, Discount Rate)\n"
        "- terminal_growth (float)\n"
        "- dilution_rate (float)\n"
        "- dividend_payout (float)\n\n"
        "Return ONLY a clean JSON object containing the overridden keys. No markdown backticks, no code blocks, no explanations. "
        "If you cannot map any variables, return an empty object {}.\n"
        "Examples:\n"
        "Input: 'What if growth rises to 12%?' -> {\"revenue_growth\": 0.12}\n"
        "Input: 'Assume margins decline by 3%.' -> If current ebit_margin is 0.31, return {\"ebit_margin\": 0.28}."
    )
    
    user_content = f"Current assumptions: {json.dumps(req.current_assumptions)}\nUser request: '{req.prompt}'"
    
    overrides = {}
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(
                f"{settings.OLLAMA_HOST}/api/generate",
                json={
                    "model": settings.OLLAMA_MODEL,
                    "prompt": f"{system_prompt}\n\n{user_content}",
                    "system": system_prompt,
                    "stream": False
                }
            )
            if response.status_code == 200:
                text = response.json().get("response", "").strip()
                if "{" in text:
                    text_clean = text[text.find("{"):text.rfind("}")+1]
                    overrides = json.loads(text_clean)
    except Exception as e:
        logger.warning(f"Ollama chat mapping failed: {e}. Falling back to regex.")

    # Fallback Regex Parsing
    prompt_lower = req.prompt.lower()
    if not overrides:
        if "growth" in prompt_lower:
            for pct in [15, 12, 10, 8, 7, 5, 4, 3, 2, 20]:
                if f"{pct}%" in prompt_lower or f"{pct} percent" in prompt_lower:
                    overrides["revenue_growth"] = pct / 100.0
                    break
        if "margin" in prompt_lower:
            for pct in [30, 25, 20, 15, 10, 35, 40]:
                if f"{pct}%" in prompt_lower or f"{pct} percent" in prompt_lower:
                    overrides["ebit_margin"] = pct / 100.0
                    break
        if "wacc" in prompt_lower or "discount" in prompt_lower:
            for pct in [10, 9, 8, 7, 11, 12]:
                if f"{pct}%" in prompt_lower or f"{pct} percent" in prompt_lower:
                    overrides["wacc"] = pct / 100.0
                    break

    # Re-run model with overridden values
    hist_financials = await fetch_historical_financials(req.ticker)
    
    new_assumptions = {**req.current_assumptions, **overrides}
    model_results = run_three_statement_model(hist_financials, new_assumptions)
    
    # Calculate consensus intrinsic value
    dcf_val = model_results["intrinsic_value"]
    provider = YahooProvider()
    try:
        price_profile = await provider.get_price(req.ticker)
        current_price = price_profile.get("price", 100.0)
    except Exception:
        current_price = 100.0

    consensus_old = run_consensus_intrinsic_value(
        dcf_val=req.current_assumptions.get("dcf_value", current_price),
        comparable_val=current_price * 1.02,
        reverse_dcf_val=current_price * 0.98,
        peg_val=current_price * 0.95,
        historical_val=current_price * 1.05
    )["intrinsic_value"]
    
    consensus_new = run_consensus_intrinsic_value(
        dcf_val=dcf_val,
        comparable_val=current_price * 1.02,
        reverse_dcf_val=current_price * 0.98,
        peg_val=current_price * 0.95,
        historical_val=current_price * 1.05
    )["intrinsic_value"]
    
    diff_pct = ((consensus_new - consensus_old) / consensus_old) * 100 if consensus_old > 0 else 0.0
    direction = "increased" if diff_pct >= 0 else "decreased"
    
    explanations = []
    for k, v in overrides.items():
        old_val = req.current_assumptions.get(k, 0.0)
        explanations.append(f"Adjusted **{k}** from `{old_val*100:.1f}%` to `{v*100:.1f}%`.")
        
    if not explanations:
        explanations.append("Could not extract any standard parameter changes. Model left unchanged.")
    else:
        explanations.append(f"This {direction} the Consensus Intrinsic Value by **{abs(diff_pct):.1f}%** (from `${consensus_old:.2f}` to `${consensus_new:.2f}`).")
        
    return {
        "status": "success",
        "overrides": overrides,
        "new_intrinsic_value": consensus_new,
        "explanation": " ".join(explanations),
        "new_assumptions": new_assumptions
    }

# 5. GET Export Excel Sheet
@router.get("/export/{model_id}")
async def export_excel(
    model_id: str,
    ticker: str = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Fetch model projections
    projections = await get_model_projections(model_id, ticker, current_user, db)
    
    # Generate openpyxl workbook
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    accent_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    # Sheet 1: Assumptions
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    
    ws_assumptions["A1"] = f"{ticker.upper()} Financial Modeling Lab Assumptions"
    ws_assumptions["A1"].font = Font(name="Calibri", size=14, bold=True)
    
    ws_assumptions["A3"] = "Parameter"
    ws_assumptions["B3"] = "Value"
    ws_assumptions["A3"].font = white_font
    ws_assumptions["B3"].font = white_font
    ws_assumptions["A3"].fill = header_fill
    ws_assumptions["B3"].fill = header_fill
    
    assumptions_dict = projections["assumptions"]
    row_idx = 4
    for key, val in assumptions_dict.items():
        ws_assumptions.cell(row=row_idx, column=1, value=key.replace("_", " ").title()).font = normal_font
        cell = ws_assumptions.cell(row=row_idx, column=2, value=val)
        cell.font = normal_font
        if isinstance(val, float) and abs(val) < 1.0:
            cell.number_format = "0.0%"
        row_idx += 1
        
    # Sheet 2: Projections (3-Statement)
    ws_proj = wb.create_sheet(title="Projections")
    ws_proj["A1"] = f"{ticker.upper()} 3-Statement Forecast Model"
    ws_proj["A1"].font = Font(name="Calibri", size=14, bold=True)
    
    years = projections["hist_years"] + projections["proj_years"]
    ws_proj.cell(row=3, column=1, value="Financial Metric").font = bold_font
    ws_proj.cell(row=3, column=1).fill = header_fill
    ws_proj.cell(row=3, column=1).font = white_font
    
    for c_idx, y in enumerate(years):
        cell = ws_proj.cell(row=3, column=c_idx + 2, value=int(y))
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right")
        
    row_labels = [
        ("Income Statement", None),
        ("Revenue", "revenue"),
        ("COGS", "cogs"),
        ("Gross Profit", "formula_gp"),
        ("EBIT", "ebit"),
        ("D&A", "dna"),
        ("EBITDA", "formula_ebitda"),
        ("Net Income", "net_income"),
        ("EPS", "eps"),
        ("Balance Sheet", None),
        ("Cash", "cash"),
        ("Working Capital", "working_capital"),
        ("Net PP&E", "net_ppe"),
        ("Total Assets", "formula_assets"),
        ("Debt", "debt"),
        ("Equity", "equity"),
        ("Total Liabilities & Equity", "formula_liabs_eq")
    ]
    
    r_idx = 4
    for label, key in row_labels:
        if key is None:
            cell = ws_proj.cell(row=r_idx, column=1, value=label)
            cell.font = Font(name="Calibri", size=11, bold=True, color="3B82F6")
            r_idx += 1
            continue
            
        ws_proj.cell(row=r_idx, column=1, value=label).font = normal_font
        
        for c_idx, y in enumerate(years):
            col_letter = chr(66 + c_idx)
            val = projections["three_statement"][y].get(key, 0.0)
            
            if key == "formula_gp":
                cell = ws_proj.cell(row=r_idx, column=c_idx + 2, value=f"={col_letter}5-{col_letter}6")
            elif key == "formula_ebitda":
                cell = ws_proj.cell(row=r_idx, column=c_idx + 2, value=f"={col_letter}8+{col_letter}9")
            elif key == "formula_assets":
                cell = ws_proj.cell(row=r_idx, column=c_idx + 2, value=f"=SUM({col_letter}14:{col_letter}16)")
            elif key == "formula_liabs_eq":
                cell = ws_proj.cell(row=r_idx, column=c_idx + 2, value=f"={col_letter}18+{col_letter}19")
            else:
                cell = ws_proj.cell(row=r_idx, column=c_idx + 2, value=val)
                
            cell.font = normal_font
            if key in ["eps"]:
                cell.number_format = "$#,##0.00"
            else:
                cell.number_format = "$#,##0"
                
        r_idx += 1

    # Sheet 3: Sensitivity Matrix
    ws_sens = wb.create_sheet(title="Sensitivity Analysis")
    ws_sens["A1"] = f"{ticker.upper()} Discount Rate (WACC) vs Revenue Growth Matrix"
    ws_sens["A1"].font = Font(name="Calibri", size=14, bold=True)
    
    sens_data = projections["sensitivity"]
    ws_sens.cell(row=3, column=1, value="WACC / Growth").font = bold_font
    ws_sens.cell(row=3, column=1).fill = header_fill
    ws_sens.cell(row=3, column=1).font = white_font
    
    for c_idx, label in enumerate(sens_data["growth_labels"]):
        cell = ws_sens.cell(row=3, column=c_idx + 2, value=label)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, wacc_label in enumerate(sens_data["wacc_labels"]):
        cell = ws_sens.cell(row=r_idx + 4, column=1, value=wacc_label)
        cell.font = bold_font
        cell.fill = header_fill
        cell.font = white_font
        
        for c_idx, val in enumerate(sens_data["matrix"][r_idx]):
            cell = ws_sens.cell(row=r_idx + 4, column=c_idx + 2, value=val)
            cell.font = normal_font
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")
            if r_idx == 2 and c_idx == 2:
                cell.fill = accent_fill
                cell.font = bold_font

    # Sheet 4: Monte Carlo Results
    ws_mc = wb.create_sheet(title="Monte Carlo")
    ws_mc["A1"] = f"{ticker.upper()} Monte Carlo Distribution Projections (10k Runs)"
    ws_mc["A1"].font = Font(name="Calibri", size=14, bold=True)
    
    mc_data = projections["monte_carlo"]
    ws_mc["A3"] = "Percentile"
    ws_mc["B3"] = "Fair Value"
    ws_mc["A3"].font = white_font
    ws_mc["B3"].font = white_font
    ws_mc["A3"].fill = header_fill
    ws_mc["B3"].fill = header_fill
    
    percentiles = [
        ("5th Percentile (Bear Case)", mc_data["p5"]),
        ("25th Percentile", mc_data["p25"]),
        ("50th Percentile (Median)", mc_data["p50"]),
        ("75th Percentile", mc_data["p75"]),
        ("95th Percentile (Bull Case)", mc_data["p95"]),
        ("Mean Simulated Value", mc_data["mean"]),
        ("Standard Deviation", mc_data["std"])
    ]
    
    for idx, (label, val) in enumerate(percentiles):
        ws_mc.cell(row=idx + 4, column=1, value=label).font = normal_font
        cell = ws_mc.cell(row=idx + 4, column=2, value=val)
        cell.font = bold_font if "50th" in label or "Mean" in label else normal_font
        cell.number_format = "$#,##0.00"
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename={ticker.upper()}_financial_model.xlsx"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@router.get("/export/pdf/{model_id}")
async def export_pdf(
    model_id: str,
    ticker: str = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Fetch model projections
    projections = await get_model_projections(model_id, ticker, current_user, db)
    
    output = io.BytesIO()
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        spaceAfter=6,
        textColor=colors.HexColor("#1F2937")
    )
    
    subtitle_style = ParagraphStyle(
        name="SubTitleStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=9,
        spaceAfter=15,
        textColor=colors.HexColor("#4B5563")
    )
    
    story.append(Paragraph(f"{ticker.upper()} Financial Projections Sheet", title_style))
    story.append(Paragraph(f"Model ID: {model_id} | User: {current_user.username}", subtitle_style))
    story.append(Spacer(1, 10))
    
    hist_years = projections["hist_years"]
    proj_years = projections["proj_years"]
    years = hist_years + proj_years
    
    table_content = [["Financial Metric"] + [str(y) for y in years]]
    
    statement = projections["three_statement"]
    metrics_mapping = [
        ("Revenue", "revenue"),
        ("EBITDA", "ebitda"),
        ("EBIT (Operating Income)", "ebit"),
        ("Tax Expense", "taxes"),
        ("CapEx", "capex"),
        ("Working Capital Change", "working_capital_change"),
    ]
    
    for label, key in metrics_mapping:
        row = [label]
        for y in years:
            year_str = str(y)
            val = statement.get(year_str, {}).get(key, 0.0)
            row.append(f"${val / 1e6:.1f}M" if abs(val) >= 1e5 else f"${val:.2f}")
        table_content.append(row)
        
    fcf_row = ["Free Cash Flow (FCF)"]
    projected_fcfs = projections["projected_fcfs"]
    
    for y in hist_years:
        year_str = str(y)
        y_data = statement.get(year_str, {})
        ebit = y_data.get("ebit", 0.0)
        tax_rate = y_data.get("tax_rate", 0.21)
        ebiat = ebit * (1 - tax_rate)
        depr = y_data.get("depreciation", 0.0)
        capex = y_data.get("capex", 0.0)
        wc = y_data.get("working_capital_change", 0.0)
        hist_fcf = ebiat + depr - capex - wc
        fcf_row.append(f"${hist_fcf / 1e6:.1f}M" if abs(hist_fcf) >= 1e5 else f"${hist_fcf:.2f}")
        
    for idx, fcf_val in enumerate(projected_fcfs):
        fcf_row.append(f"${fcf_val / 1e6:.1f}M" if abs(fcf_val) >= 1e5 else f"${fcf_val:.2f}")
        
    table_content.append(fcf_row)
    
    disc_fcf_row = ["Discounted FCF"]
    for y in hist_years:
        disc_fcf_row.append("-")
    for idx, d_fcf in enumerate(projections["discounted_fcfs"]):
        disc_fcf_row.append(f"${d_fcf / 1e6:.1f}M" if abs(d_fcf) >= 1e5 else f"${d_fcf:.2f}")
    table_content.append(disc_fcf_row)
    
    t = Table(table_content)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F2937")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-2), (-1,-2), colors.HexColor("#EFF6FF")),
    ]))
    
    story.append(t)
    story.append(Spacer(1, 20))
    story.append(Paragraph("Valuation Model Metrics Summary", styles["Heading3"]))
    story.append(Spacer(1, 6))
    
    wacc = projections["assumptions"]["wacc"]
    terminal_growth = projections["assumptions"]["terminal_growth"]
    intrinsic_val = projections["intrinsic_value"]
    current_price = projections["current_price"]
    mos = projections["margin_of_safety"]
    
    summary_text = (
        f"<b>Weighted Cost of Capital (WACC):</b> {wacc*100:.2f}% &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Terminal Growth Rate:</b> {terminal_growth*100:.2f}%<br/>"
        f"<b>Calculated Intrinsic Fair Value:</b> ${intrinsic_val:.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Current Price:</b> ${current_price:.2f} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Margin of Safety:</b> {mos*100:.2f}%"
    )
    story.append(Paragraph(summary_text, styles["Normal"]))
    
    doc.build(story)
    output.seek(0)
    
    headers_resp = {
        "Content-Disposition": f"attachment; filename={ticker.upper()}_financial_model.pdf"
    }
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=headers_resp
    )
