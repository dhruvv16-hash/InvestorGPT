from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.database.db import get_db
from app.models.models import PortfolioHolding
from app.providers.market.yahoo_provider import YahooProvider
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io
import httpx
import time
from datetime import datetime, timezone
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from app.dependencies import get_current_user
from app.models.models import User

logger = logging.getLogger("investorgpt.routes_portfolio")
router = APIRouter(prefix="/portfolio", tags=["Portfolio"])

class HoldingAddRequest(BaseModel):
    user_id: str
    ticker: str
    shares: float
    price: float

EXCHANGE_CACHE = {}
CACHE_EXPIRY = {}

def get_ticker_currency(ticker: str) -> str:
    t = ticker.upper().strip()
    if t.endswith(".NS") or t.endswith(".BO") or "PW" in t or "WALLAH" in t:
        return "INR"
    if t.endswith(".L"):
        return "GBP"
    if t.endswith(".PA") or t.endswith(".DE"):
        return "EUR"
    if t.endswith(".T"):
        return "JPY"
    return "USD"

async def get_exchange_rate(from_curr: str, to_curr: str) -> float:
    from_curr = from_curr.upper().strip()
    to_curr = to_curr.upper().strip()
    if from_curr == to_curr:
        return 1.0
        
    cache_key = f"{from_curr}_{to_curr}"
    now = time.time()
    if cache_key in EXCHANGE_CACHE and now - CACHE_EXPIRY.get(cache_key, 0) < 3600:
        return EXCHANGE_CACHE[cache_key]
        
    try:
        url = f"https://open.er-api.com/v6/latest/{from_curr}"
        async with httpx.AsyncClient(timeout=5.0) as client:
            res = await client.get(url)
            if res.status_code == 200:
                data = res.json()
                rates = data.get("rates", {})
                rate = rates.get(to_curr)
                if rate:
                    val = float(rate)
                    EXCHANGE_CACHE[cache_key] = val
                    CACHE_EXPIRY[cache_key] = now
                    # Cache the reverse too
                    rev_key = f"{to_curr}_{from_curr}"
                    EXCHANGE_CACHE[rev_key] = 1.0 / val
                    CACHE_EXPIRY[rev_key] = now
                    return val
    except Exception as e:
        logger.error(f"Failed to fetch exchange rate from {from_curr} to {to_curr}: {e}")
        
    fallbacks = {
        ("USD", "INR"): 83.5,
        ("INR", "USD"): 1.0 / 83.5,
        ("USD", "EUR"): 0.92,
        ("EUR", "USD"): 1.0 / 0.92,
        ("USD", "GBP"): 0.79,
        ("GBP", "USD"): 1.0 / 0.79
    }
    return fallbacks.get((from_curr, to_curr), 1.0)

def get_currency_symbol(currency: str) -> str:
    upper = currency.upper().strip()
    if upper == "INR": return "₹"
    if upper == "EUR": return "€"
    if upper == "GBP": return "£"
    if upper == "JPY": return "¥"
    return "$"

@router.get("")
async def get_portfolio(
    preferred_currency: str = Query("USD"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    holdings = db.query(PortfolioHolding).filter(PortfolioHolding.user_id == current_user.id).all()
    pref_curr = preferred_currency.upper().strip()
    
    total_cost = 0.0
    total_value = 0.0
    
    holdings_data = []
    provider = YahooProvider()
    
    for h in holdings:
        ticker = h.ticker.upper().strip()
        shares = float(h.shares)
        
        # Native details
        native_currency = get_ticker_currency(ticker)
        avg_price_native = float(h.avg_buy_price)
        
        # Fetch current price dynamically (in native currency)
        current_price_native = avg_price_native
        try:
            price_info = await provider.get_price(ticker)
            current_price_native = price_info["price"]
        except Exception as e:
            logger.warning(f"Failed to fetch live price for {ticker}: {e}")
            
        # Get conversion rate to preferred display currency
        rate = await get_exchange_rate(native_currency, pref_curr)
        
        # Calculate cost and value in preferred currency
        cost_pref = avg_price_native * shares * rate
        val_pref = current_price_native * shares * rate
        
        total_cost += cost_pref
        total_value += val_pref
        
        pnl_pref = val_pref - cost_pref
        pnl_pct = (pnl_pref / cost_pref * 100.0) if cost_pref > 0 else 0.0
        
        holdings_data.append({
            "id": h.id,
            "ticker": ticker,
            "shares": shares,
            "native_currency": native_currency,
            "avg_buy_price": avg_price_native * rate,
            "current_price": current_price_native * rate,
            "avg_buy_price_native": avg_price_native,
            "current_price_native": current_price_native,
            "cost": cost_pref,
            "value": val_pref,
            "pnl": pnl_pref,
            "pnl_pct": pnl_pct
        })
        
    for h_data in holdings_data:
        h_data["weight_pct"] = (h_data["value"] / total_value * 100.0) if total_value > 0 else 0.0
        
    total_pnl = total_value - total_cost
    total_pnl_pct = (total_pnl / total_cost * 100.0) if total_cost > 0 else 0.0
    
    return {
        "holdings": holdings_data,
        "preferred_currency": pref_curr,
        "currency_symbol": get_currency_symbol(pref_curr),
        "summary": {
            "total_cost": total_cost,
            "total_value": total_value,
            "total_pnl": total_pnl,
            "total_pnl_pct": total_pnl_pct
        }
    }

@router.post("/add")
def add_holding(
    req: HoldingAddRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ticker = req.ticker.upper().strip()
    if not ticker or req.shares <= 0 or req.price <= 0:
        raise HTTPException(status_code=400, detail="Invalid shares or buy price.")
        
    existing = db.query(PortfolioHolding).filter(
        PortfolioHolding.user_id == current_user.id,
        PortfolioHolding.ticker == ticker
    ).first()
    
    if existing:
        old_shares = float(existing.shares)
        old_avg = float(existing.avg_buy_price)
        
        new_shares = old_shares + req.shares
        new_avg = (old_shares * old_avg + req.shares * req.price) / new_shares
        
        existing.shares = new_shares
        existing.avg_buy_price = new_avg
        db.commit()
        db.refresh(existing)
        return {"status": "success", "action": "updated", "id": existing.id}
    else:
        new_h = PortfolioHolding(
            user_id=current_user.id,
            ticker=ticker,
            shares=req.shares,
            avg_buy_price=req.price
        )
        db.add(new_h)
        db.commit()
        db.refresh(new_h)
        return {"status": "success", "action": "created", "id": new_h.id}

@router.delete("/remove/{holding_id}")
def remove_holding(
    holding_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    holding = db.query(PortfolioHolding).filter(
        PortfolioHolding.id == holding_id,
        PortfolioHolding.user_id == current_user.id
    ).first()
    if not holding:
        raise HTTPException(status_code=404, detail="Holding not found or not owned by user.")
    db.delete(holding)
    db.commit()
    return {"status": "success"}

@router.get("/export/excel")
async def export_portfolio_excel(
    preferred_currency: str = Query("USD"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    data = await get_portfolio(preferred_currency, current_user, db)
    holdings = data["holdings"]
    summary = data["summary"]
    pref_curr = data["preferred_currency"]
    sym = data["currency_symbol"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Holdings"
    
    ws["A1"] = f"InvestorGPT Portfolio Holdings Report ({pref_curr})"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A2"] = f"User: {current_user.username} | Display Currency: {pref_curr} ({sym})"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    
    headers = [
        "Ticker", "Shares", "Avg Entry Price", "Current Price", 
        f"Total Cost ({sym})", f"Total Value ({sym})", f"Gain / Loss ({sym})", "Gain / Loss (%)", "Weight (%)"
    ]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    for col_idx, h in enumerate(headers):
        cell = ws.cell(row=4, column=col_idx + 1, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 0 else "left")
        
    num_fmt = f'"{sym}"#,##0.00'
    
    row_num = 5
    for h in holdings:
        ws.cell(row=row_num, column=1, value=h["ticker"]).font = bold_font
        ws.cell(row=row_num, column=2, value=h["shares"]).font = normal_font
        
        c3 = ws.cell(row=row_num, column=3, value=h["avg_buy_price"])
        c3.number_format = num_fmt
        c3.font = normal_font
        
        c4 = ws.cell(row=row_num, column=4, value=h["current_price"])
        c4.number_format = num_fmt
        c4.font = normal_font
        
        c5 = ws.cell(row=row_num, column=5, value=h["cost"])
        c5.number_format = num_fmt
        c5.font = normal_font
        
        c6 = ws.cell(row=row_num, column=6, value=h["value"])
        c6.number_format = num_fmt
        c6.font = normal_font
        
        c7 = ws.cell(row=row_num, column=7, value=h["pnl"])
        c7.number_format = num_fmt
        c7.font = normal_font
        
        c8 = ws.cell(row=row_num, column=8, value=h["pnl_pct"] / 100.0)
        c8.number_format = "0.0%"
        c8.font = normal_font
        
        c9 = ws.cell(row=row_num, column=9, value=h["weight_pct"] / 100.0)
        c9.number_format = "0.0%"
        c9.font = normal_font
        
        row_num += 1
        
    ws.cell(row=row_num, column=1, value="TOTAL").font = bold_font
    ws.cell(row=row_num, column=2, value="").font = bold_font
    ws.cell(row=row_num, column=3, value="").font = bold_font
    ws.cell(row=row_num, column=4, value="").font = bold_font
    
    t_cost = ws.cell(row=row_num, column=5, value=summary["total_cost"])
    t_cost.number_format = num_fmt
    t_cost.font = bold_font
    
    t_val = ws.cell(row=row_num, column=6, value=summary["total_value"])
    t_val.number_format = num_fmt
    t_val.font = bold_font
    
    t_pnl = ws.cell(row=row_num, column=7, value=summary["total_pnl"])
    t_pnl.number_format = num_fmt
    t_pnl.font = bold_font
    
    t_pnl_pct = ws.cell(row=row_num, column=8, value=summary["total_pnl_pct"] / 100.0)
    t_pnl_pct.number_format = "0.0%"
    t_pnl_pct.font = bold_font
    
    ws.cell(row=row_num, column=9, value=1.0).font = bold_font
    ws.cell(row=row_num, column=9).number_format = "0.0%"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_resp = {
        "Content-Disposition": f"attachment; filename=portfolio_report_{pref_curr.lower()}.xlsx"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp
    )

@router.get("/export/pdf")
async def export_portfolio_pdf(
    preferred_currency: str = Query("USD"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    data = await get_portfolio(preferred_currency, current_user, db)
    holdings = data["holdings"]
    summary = data["summary"]
    pref_curr = data["preferred_currency"]
    sym = data["currency_symbol"]
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        spaceAfter=6,
        textColor=colors.HexColor("#1F2937")
    )
    
    subtitle_style = ParagraphStyle(
        name="SubTitleStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Oblique",
        fontSize=9,
        spaceAfter=20,
        textColor=colors.HexColor("#4B5563")
    )
    
    story.append(Paragraph(f"InvestorGPT Portfolio Holdings Report ({pref_curr})", title_style))
    story.append(Paragraph(f"User: {current_user.username} | Display Currency: {pref_curr} ({sym})", subtitle_style))
    story.append(Spacer(1, 10))
    
    table_content = [[
        "Ticker", "Shares", "Avg Entry", "Current Price", 
        f"Cost ({sym})", f"Value ({sym})", f"P&L ({sym})", "P&L (%)", "Weight"
    ]]
    
    for h in holdings:
        table_content.append([
            h["ticker"],
            f"{h['shares']:.2f}",
            f"{sym}{h['avg_buy_price']:.2f}",
            f"{sym}{h['current_price']:.2f}",
            f"{sym}{h['cost']:.2f}",
            f"{sym}{h['value']:.2f}",
            f"{sym}{h['pnl']:.2f}",
            f"{h['pnl_pct']:.1f}%",
            f"{h['weight_pct']:.1f}%"
        ])
        
    table_content.append([
        "TOTAL",
        "",
        "",
        "",
        f"{sym}{summary['total_cost']:.2f}",
        f"{sym}{summary['total_value']:.2f}",
        f"{sym}{summary['total_pnl']:.2f}",
        f"{summary['total_pnl_pct']:.1f}%",
        "100.0%"
    ])
    
    t = Table(table_content)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F2937")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor("#F9FAFB"), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#F3F4F6")),
        ('TOPPADDING', (0,-1), (-1,-1), 8),
        ('BOTTOMPADDING', (0,-1), (-1,-1), 8),
    ]))
    
    story.append(t)
    doc.build(story)
    output.seek(0)
    
    headers_resp = {
        "Content-Disposition": f"attachment; filename=portfolio_report_{pref_curr.lower()}.pdf"
    }
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=headers_resp
    )

@router.get("/optimize")
async def optimize_portfolio_holdings(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    data = await get_portfolio("USD", current_user, db)
    holdings = data.get("holdings", [])
    if not holdings:
        raise HTTPException(status_code=400, detail="Cannot optimize an empty portfolio. Please add holdings first.")
    
    from app.engines.portfolio_optimization import PortfolioOptimizationEngine
    engine = PortfolioOptimizationEngine()
    result = await engine.optimize_portfolio(holdings)
    return result
